#This script reads through a Voyager import log and outputs duplicate bib IDs as well as the IDs of bibs, mfhds, and items created.

#import regular expressions and openpyxl
import re
import openpyxl

# prompt for file names
fname = input("Enter input file, including extension: ")
fout = input("Enter output file, without extension: ")
fh = open(fname, "r")

# set up lists
duplicates = [["Duplicate Bib ID"]]
bibs = [["Bib ID"]]
mfhds = [["MFHD ID"]]
items = [["Item ID"]]

# create and open workbook with two sheets
wb1=openpyxl.Workbook()
ws1=wb1.active
ws1.title = "Duplicate Bibs"
ws2 = wb1.create_sheet(index=1, title="IDs Added")

# read through file, extract the line after the line starting with BibID & rank and write to lists
with fh as f:
    lines = f.readlines()
    n_lines = len(lines)
    for i, line in enumerate (lines) :
        line = line.rstrip()
        if line.startswith("	BibID & rank") and \
        n_lines > i + 2 and lines[i + 2].startswith("") :
            bibline = re.findall(r'\d+\s-\s', lines[i+1])
            dupeid = re.findall(r'\d+', str(bibline))
            duplicates.append(dupeid)
        elif line.startswith("	Adding Bib") :
            line = re.findall(r'\d+',str(line))
            bibs.append(line)
        elif line.startswith("MFHD_ID ") :
            line = re.findall(r'\d+',str(line))
            mfhds.append(line)
        elif line.startswith("ITEM_ID ") :
            line = re.findall(r'\d+',str(line))
            items.append(line)
        else :
            continue

# write the lists to columns in the spreadsheet and save
for row in duplicates:
    ws1.append(row)
for r in range(0,len(bibs)):
    ws2.cell(row=r+1,column=1).value=bibs[r][0]
for r in range(0,len(mfhds)):
    ws2.cell(row=r+1,column=2).value=mfhds[r][0]
for r in range(0,len(items)):
    ws2.cell(row=r+1,column=3).value=items[r][0]
wb1.save(fout + ".xlsx")
